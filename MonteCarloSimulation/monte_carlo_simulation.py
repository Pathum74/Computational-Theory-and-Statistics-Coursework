import random
import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

def estimate_pi(num_points):
    inside_circle = 0
    for _ in range(num_points):
        x = random.uniform(-1, 1)
        y = random.uniform(-1, 1)
        if x**2 + y**2 <= 1:
            inside_circle += 1
    return 4 * (inside_circle / num_points)

if __name__ == "__main__":
    num_points_list = [1000, 10000, 100000, 1000000]
    num_repeats = 10
    results = []

    for num_points in num_points_list:
        estimates = []
        for _ in range(num_repeats):
            pi_estimate = estimate_pi(num_points)
            estimates.append(pi_estimate)
        results.append({
            "Num Points": num_points,
            "Mean π": sum(estimates) / num_repeats,
            "Mode π": max(set(estimates), key=estimates.count),
            "Estimates": estimates
        })

with open("pi_estimates.csv", "w", newline="", encoding="utf-8") as file:
    writer = csv.writer(file)
    writer.writerow(["Num Points", "Mean pi", "Mode pi", "Estimates"])
    for result in results:
        writer.writerow([result["Num Points"], result["Mean π"], result["Mode π"], result["Estimates"]])


    print("Results saved to pi_estimates.csv")

csv_file = 'pi_estimates.csv'
excel_file = 'pi_estimates.xlsx'

try:
    data = pd.read_csv(csv_file)

    data.to_excel(excel_file, index=False, engine='openpyxl')

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)

    for col_idx, col_cells in enumerate(sheet.iter_cols(), start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in col_cells:
            cell.border = thin_border
            if cell.row == 1:
                cell.font = bold_font
                cell.alignment = center_alignment

            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        if column_letter == 'C' and sheet['C1'].value == 'Estimates':
            sheet.column_dimensions[column_letter].width = 90
        else:
            sheet.column_dimensions[column_letter].width = max_length + 2

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
        sheet.row_dimensions[row[0].row].height = 15

    workbook.save(excel_file)
    print(f"Data successfully saved and formatted in {excel_file}")

except FileNotFoundError:
    print(f"The file {csv_file} was not found. Ensure it's in the same directory as this script.")
except Exception as e:
    print(f"An error occurred: {e}")

