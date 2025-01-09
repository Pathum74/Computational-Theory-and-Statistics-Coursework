import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side

num_simulations = 100000

children = np.random.randint(0, 2, (num_simulations, 2))
at_least_one_girl = np.any(children == 1, axis=1)

both_girls = np.all(children == 1, axis=1)

probability = np.sum(both_girls & at_least_one_girl) / np.sum(at_least_one_girl)

print(f"Probability that all children are girls given at least one is a girl: {probability:.5f}")

num_simulations = 100000

dice_rolls = np.random.randint(1, 7, (num_simulations, 10))

sum_equals_30 = np.sum(dice_rolls, axis=1) == 30

probability = np.sum(sum_equals_30) / num_simulations

print(f"Probability that the sum of 10 dice rolls equals 30: {probability:.5f}")

# Create a DataFrame for the results
results = pd.DataFrame({
    "Simulation": ["Probability of all girls given at least one girl",
                   "Probability of sum of 10 dice rolls equals 30"],
    "Probability": [np.sum(both_girls & at_least_one_girl) / np.sum(at_least_one_girl),
                    np.sum(sum_equals_30) / num_simulations]
})

# Save to CSV
results.to_csv("simulation_results.csv", index=False)

# Create Excel file with formatting
wb = Workbook()
ws = wb.active
ws.title = "Results"

# Add headers
headers = list(results.columns)
ws.append(headers)

# Style headers
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data
for row in dataframe_to_rows(results, index=False, header=False):
    ws.append(row)

# Style data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get column letter
    for cell in col:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Add borders to table
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.border = thin_border

# Save the file
wb.save("simulation_results.xlsx")

print("Results saved to 'simulation_results.xlsx'")
